"""
One-time extraction of the pricing catalog from the studio's Excel workbook
into pricing_data.json, which becomes the app's source of truth from then on.

Reads:
  - Okapics sheet: flat catalog of {Product, Category, Item} -> cost/profit
    parameters (material side + labor/"effort" side, already resolved to
    currency by the workbook itself, including array-formula rows).
  - Calculator sheet: for every in-scope component row, detects whether its
    cost scales with the job's perimeter or its area by pattern-matching
    the row's own formula text (never hardcoded per category name), and
    whether the row is a fixed single component (toggled Yes/No) or a
    dropdown offering every item in an Okapics category (confirmed via the
    sheet's own data-validation rules, not guessed).

Ignores 'Price Check' and 'backup' entirely, per instructions. Also
excludes (see EXCLUDED_FEATURES): the fully-custom wood-profile builder
(row 22), Crate (rows 34-36, which use a 3D X/Y/Z box formula, not
height/width), and the CNC section (rows 38-46, which uses a different
quantity model and hasn't been verified against the workbook).
"""
import json
import re
import sys
from pathlib import Path

import openpyxl
from oletools.olevba import VBA_Parser

SOURCE_XLSM = Path(__file__).resolve().parent.parent / "PrintHouse Pricing Calc v8_3 - 09_03_2025.xlsm"
OUTPUT_JSON = Path(__file__).resolve().parent / "pricing_data.json"

# Okapics column indices (1-based), from the sheet's own header row.
OK_PRODUCT, OK_CATEGORY, OK_ITEM = 1, 2, 3
OK_FIXED_MAT, OK_VAR_MAT, OK_MAT_INDIRECT, OK_MAT_RISK, OK_MAT_PROFIT = 4, 5, 6, 7, 9
OK_VAR_EFFORT, OK_FIXED_EFFORT, OK_MIN_EFFORT = 14, 15, 16
OK_EFFORT_INDIRECT, OK_EFFORT_RISK, OK_EFFORT_PROFIT = 17, 18, 20

# Calculator columns (1-based).
CAL_PRODUCT, CAL_CATEGORY, CAL_SUBCATEGORY, CAL_ITEM, CAL_ADD, CAL_H = 3, 4, 5, 6, 7, 8

PERIMETER_RE = re.compile(r"\$[A-Z]+\$\d+\+\$[A-Z]+\$\d+\)\*2/100")
AREA_RE = re.compile(r"\$[A-Z]+\$\d+\*\$[A-Z]+\$\d+\)/10000")

# Rows 11-32 are multiplied by order Quantity in the workbook's summary
# formula (Calculator!W16); services are not. See pricing_engine.py for
# the full reproduction of that formula.
QUANTITY_SCALED_ROWS = range(11, 33)

# Confirmed via Calculator's own Data Validation rules (openpyxl
# ws.data_validations): these rows have a dropdown on column F offering
# every item of an Okapics category. Every other in-scope row is a fixed,
# single named component that's simply toggled on/off (or, for Services,
# multiplied by an hours quantity).
DROPDOWN_ROWS = {11, 23, 24, 25, 26, 27, 28, 29, 30, 31}

# In-scope rows. Excludes: row 7 "Testing Paper" (not in the Okapics
# catalog at all), row 22 "Custom Profile" (custom wood builder), rows
# 34-36 "Crate" (3D X/Y/Z formula), rows 38-46 "CNC" (different quantity
# model, unverified).
IN_SCOPE_ROWS = [5, 6, 8, 9] + [11] + list(range(12, 22)) + [23, 24, 25, 26, 27, 28, 29, 30, 31, 32]

SERVICES_ROWS = {5, 6, 8, 9}

EXCLUDED_FEATURES = [
    "Row 7 'Testing Paper' has no matching entry in the Okapics catalog and is not priced.",
    "Custom Profile (row 22): fully custom wood-profile builder (pick wood species/thickness/alignment) is not supported yet.",
    "Crate (rows 34-36): uses a 3D box (X/Y/Z) formula rather than height/width and is not supported yet.",
    "CNC section (rows 38-46): uses a different quantity model (decimal input, not Yes/No) and has not been verified against the workbook; not supported yet.",
]


def formula_text(cell_value):
    return cell_value.text if hasattr(cell_value, "text") else cell_value


def slugify(text):
    return re.sub(r"[^a-z0-9]+", "_", str(text).lower()).strip("_")


def load_okapics_catalog(wb_values):
    """
    A handful of item names appear more than once in the sheet under
    different categories (e.g. 'Double Sided tape' exists under both
    'Drawing Methods' and 'Glueing', with different costs). The
    Calculator's actual formula does `MATCH($F, Okapics!C:C, 0)`, which
    returns the FIRST matching row — so on a duplicate name, we must keep
    the first occurrence too, not the last, to stay faithful to what
    Excel actually computes.
    """
    ws = wb_values["Okapics"]
    items = {}
    for r in range(3, ws.max_row + 1):
        item = ws.cell(row=r, column=OK_ITEM).value
        if not item:
            continue
        key = item.strip() if isinstance(item, str) else item
        if key in items:
            continue
        items[key] = {
            "product": ws.cell(row=r, column=OK_PRODUCT).value,
            "category": ws.cell(row=r, column=OK_CATEGORY).value,
            "item": key,
            "fixed_material_cost": ws.cell(row=r, column=OK_FIXED_MAT).value or 0,
            "var_material_cost": ws.cell(row=r, column=OK_VAR_MAT).value or 0,
            "material_indirect": ws.cell(row=r, column=OK_MAT_INDIRECT).value or 0,
            "material_risk": ws.cell(row=r, column=OK_MAT_RISK).value or 0,
            "material_profit": ws.cell(row=r, column=OK_MAT_PROFIT).value or 0,
            "var_effort_cost": ws.cell(row=r, column=OK_VAR_EFFORT).value or 0,
            "fixed_effort_cost": ws.cell(row=r, column=OK_FIXED_EFFORT).value or 0,
            "min_effort_cost": ws.cell(row=r, column=OK_MIN_EFFORT).value or 0,
            "effort_indirect": ws.cell(row=r, column=OK_EFFORT_INDIRECT).value or 0,
            "effort_risk": ws.cell(row=r, column=OK_EFFORT_RISK).value or 0,
            "effort_profit": ws.cell(row=r, column=OK_EFFORT_PROFIT).value or 0,
        }
    return items


def detect_size_mode(formula_str, row_num):
    if formula_str is None:
        return None
    flat = "".join(formula_str.split())
    if PERIMETER_RE.search(flat):
        return "perimeter"
    if AREA_RE.search(flat):
        return "area"
    if flat.startswith(f"=G{row_num}*") or flat.startswith(f"=IF(G{row_num}>0,"):
        return "manual_hours"
    return "unknown"


def item_public_fields(rec):
    return {k: v for k, v in rec.items() if k not in ("product", "category", "item")}


# The workbook has 15 macro "buttons" (VBA Sub procedures in Module1.bas)
# that reset the Calculator and then flip a specific bundle of rows to
# "Yes" with specific default items — these are the studio's real-world
# framing "packages" (Box, Float, Aluminium, Canvas, Facemount...). We
# parse them straight out of the macro source so they can never drift
# from what the buttons in Excel actually do.
NON_PRESET_MACROS = {"Button68_Click", "Macro1", "ded", "ResetAll"}

SUB_RE = re.compile(r"Sub\s+(\w+)\s*\(\)(.*?)End Sub", re.S)
YES_RANGE_RE = re.compile(r'Range\("([^"]+)"\)\.Value\s*=\s*"Yes"')
CELL_ASSIGN_RE = re.compile(r'Range\("([EF])(\d+)"\)\.Value\s*=\s*"([^"]*)"')

PRESET_LABELS = {
    "Box": "Box Frame",
    "Box_Mount_Print": "Box Frame + Mount + Print",
    "Box_Drawing": "Box Frame + Drawing",
    "LightBox": "Light Box",
    "Float": "Float Frame (no print — for an existing piece)",
    "Float_Chromalux": "Float Frame + Chromalux",
    "Canvas": "Canvas Stretch",
    "Facemount": "Facemount",
    "Facemount_Float": "Float Frame + Print, Acrylic Facemount",
    "Kapa_Float": "Float Frame + Print (default — Kapa back)",
    "Print_on_dbond": "Print + Dibond",
    "Aluminium_Drawing": "Aluminium Frame + Drawing",
    "Aluminium_Print": "Aluminium Frame + Print",
    "Aluminium": "Aluminium Frame",
    "Chromalux": "Chromalux",
}

# Manual corrections to what the raw VBA macros default to, confirmed by
# the studio as out of date. Keyed by macro name -> {slot_key: item_name}.
# The macros themselves still say "Float Dibond/Kapa 3/6 simple"; the
# studio's actual default float profile is "3/4.5". Likewise the macros
# say "Double Sided tape" for drawing; the studio's actual default is
# "Hinjes with japanies paper" (Hinges with Japanese paper).
PRESET_COMPONENT_OVERRIDES = {
    "Float": {"row23_profile_preset": "Float Dibond/Kapa 3/4.5 simple"},
    "Float_Chromalux": {"row23_profile_preset": "Float Dibond/Kapa 3/4.5 simple"},
    "Facemount_Float": {"row23_profile_preset": "Float Dibond/Kapa 3/4.5 simple"},
    "Kapa_Float": {"row23_profile_preset": "Float Dibond/Kapa 3/4.5 simple"},
    "Box_Drawing": {"row28_drawing": "Hinjes with japanies paper"},
    "Aluminium_Drawing": {"row28_drawing": "Hinjes with japanies paper"},
}


def extract_presets(source_xlsm, slots):
    row_to_slot = {slot["row"]: slot for slot in slots}

    vba = VBA_Parser(str(source_xlsm))
    module_source = None
    for (_, _, vba_filename, code) in vba.extract_macros():
        if vba_filename == "Module1.bas":
            module_source = code
            break
    if module_source is None:
        return []

    presets = []
    for match in SUB_RE.finditer(module_source):
        name, body = match.group(1), match.group(2)
        if name in NON_PRESET_MACROS:
            continue

        enabled_rows = []
        for cell_list in YES_RANGE_RE.findall(body):
            for ref in cell_list.split(","):
                m = re.fullmatch(r"G(\d+)", ref.strip())
                if m:
                    enabled_rows.append(int(m.group(1)))

        overrides = {}  # row -> {"E": val, "F": val}
        for col, row_str, value in CELL_ASSIGN_RE.findall(body):
            overrides.setdefault(int(row_str), {})[col] = value

        manual_overrides = PRESET_COMPONENT_OVERRIDES.get(name, {})

        components = []
        skipped_rows = []
        for row in enabled_rows:
            slot = row_to_slot.get(row)
            if slot is None:
                skipped_rows.append(row)
                continue
            f_override = overrides.get(row, {}).get("F")
            item_name = f_override.strip() if f_override else slot["default_item"]
            if item_name not in slot["items"]:
                # Fall back to whatever this slot actually offers, rather
                # than dropping the component silently.
                item_name = slot["default_item"]
            item_name = manual_overrides.get(slot["key"], item_name)
            components.append({"slot_key": slot["key"], "item_name": item_name})

        presets.append({
            "key": slugify(name),
            "macro_name": name,
            "label": PRESET_LABELS.get(name, name.replace("_", " ")),
            "components": components,
            "unsupported_rows": skipped_rows,
        })

    return presets


def extract():
    if not SOURCE_XLSM.exists():
        sys.exit(f"Source workbook not found: {SOURCE_XLSM}")

    wb_values = openpyxl.load_workbook(SOURCE_XLSM, data_only=True, keep_vba=True)
    wb_formulas = openpyxl.load_workbook(SOURCE_XLSM, data_only=False, keep_vba=True)

    okapics = load_okapics_catalog(wb_values)

    calc_v = wb_values["Calculator"]
    calc_f = wb_formulas["Calculator"]

    salary_per_hour = calc_v["W45"].value or 90
    # The source workbook hardcodes 17% VAT (Calculator!Y11/Y16 use *1.17).
    # Updated to 18% per the studio's current rate — kept here (not read
    # from the sheet) so re-running this script doesn't silently revert it.
    vat_rate = 0.18

    slots = []
    skipped = []

    product, category, subcategory = None, None, None
    for r in range(5, 33):
        c_product = calc_v.cell(row=r, column=CAL_PRODUCT).value
        c_category = calc_v.cell(row=r, column=CAL_CATEGORY).value
        c_sub = calc_v.cell(row=r, column=CAL_SUBCATEGORY).value
        if c_product:
            product = c_product
            category = None
            subcategory = None
        if c_category:
            category = c_category
            subcategory = None
        if c_sub:
            subcategory = c_sub

        if r not in IN_SCOPE_ROWS:
            continue

        item_name = calc_v.cell(row=r, column=CAL_ITEM).value
        if not item_name:
            continue
        item_name = item_name.strip() if isinstance(item_name, str) else item_name

        catalog_entry = okapics.get(item_name)
        if catalog_entry is None:
            skipped.append((r, item_name, "not found in Okapics catalog"))
            continue

        h_formula = formula_text(calc_f.cell(row=r, column=CAL_H).value)
        size_mode = detect_size_mode(h_formula, r)
        if r in SERVICES_ROWS:
            size_mode = "manual_hours"
        if size_mode in (None, "unknown"):
            skipped.append((r, item_name, f"could not determine size mode from formula"))
            continue

        label = item_name if r not in DROPDOWN_ROWS else (category or subcategory or item_name)
        slot_key = f"row{r}_{slugify(label)}"

        if r in DROPDOWN_ROWS:
            ok_product = catalog_entry["product"]
            ok_category = catalog_entry["category"]
            items = {
                name: item_public_fields(rec)
                for name, rec in okapics.items()
                if rec["product"] == ok_product and rec["category"] == ok_category
            }
            kind = "dropdown"
        else:
            items = {item_name: item_public_fields(catalog_entry)}
            kind = "fixed"

        slots.append({
            "key": slot_key,
            "row": r,
            "kind": kind,  # "dropdown" (pick one item) or "fixed" (this exact item, toggled on/off)
            "product": product,
            "category": category,
            "subcategory": subcategory,
            "label": label,
            "size_mode": size_mode,  # "perimeter", "area", or "manual_hours"
            "scales_with_quantity": r in QUANTITY_SCALED_ROWS,
            "default_item": item_name,
            "items": items,
        })

    presets = extract_presets(SOURCE_XLSM, slots)

    data = {
        "source_file": SOURCE_XLSM.name,
        "salary_per_hour": salary_per_hour,
        "vat_rate": vat_rate,
        "slots": slots,
        "presets": presets,
        "excluded_features": EXCLUDED_FEATURES,
    }

    OUTPUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Wrote {OUTPUT_JSON} with {len(slots)} slots and {len(presets)} presets.")
    for slot in slots:
        print(f"  - {slot['key']} [{slot['kind']}] size_mode={slot['size_mode']} "
              f"items={len(slot['items'])} qty_scaled={slot['scales_with_quantity']}")
    if skipped:
        print("\nSkipped rows:")
        for r, name, reason in skipped:
            print(f"  - row {r} ({name!r}): {reason}")

    print("\nPresets:")
    for preset in presets:
        note = f"  [unsupported rows: {preset['unsupported_rows']}]" if preset["unsupported_rows"] else ""
        print(f"  - {preset['key']} ({preset['macro_name']}): {len(preset['components'])} components{note}")


if __name__ == "__main__":
    extract()
